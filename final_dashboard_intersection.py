import re
from io import BytesIO
import openpyxl
import pandas as pd
import plotly.express as px
import streamlit as st

AGENCY_COLUMNS = {"IPSOS", "KANTAR"}
NON_RESPONSE_ROWS = {
    "UNWEIGHTED SAMPLE",
    "WEIGHTED SAMPLE",
    "MEAN",
    "MEDIAN",
    "MODE",
    "SD",
    "SE",
    "STD",
    "STANDARD ERROR",
    "STANDARD DEVIATION",
}
COLOR_PALETTES = {
    "Default": ["#4C78A8", "#F58518", "#54A24B", "#E45756", "#72B7B2", "#B279A2"],
    "Blue": ["#1f77b4", "#6baed6", "#9ecae1", "#c6dbef", "#08306b"],
    "Green": ["#2ca02c", "#74c476", "#a1d99b", "#c7e9c0", "#005a32"],
    "Warm": ["#e45756", "#f58518", "#ffbf79", "#b30000", "#7f2704"],
    "Pastel": ["#9ecae1", "#fdd0a2", "#c7e9c0", "#fcbba1", "#dadaeb"],
    "Grayscale": ["#252525", "#636363", "#969696", "#bdbdbd", "#d9d9d9"],
}
CHART_SIZES = {
    "Small": (9, 5),
    "Medium": (12, 6),
    "Large": (15, 8),
}
EXCEL_CHART_SIZES = {
    "Small": (18, 9),
    "Medium": (24, 12),
    "Large": (30, 16),
}


st.set_page_config(
    page_title="Maruti Survey Dashboard",
    layout="wide",
    initial_sidebar_state="expanded",
)


def to_number(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def clean_title(text: str) -> str:
    cleaned = str(text).strip()
    cleaned = re.sub(r'\s*\([^)]*\)\s*', ' ', cleaned)
    cleaned = re.sub(r'^\s*[a-z0-9_.-]*\s*\.\s*', '', cleaned, flags=re.IGNORECASE)
    return re.sub(r'\s+', ' ', cleaned).strip()


def safe_filename(text: str) -> str:
    cleaned = "".join(character if character.isalnum() else "_" for character in str(text).lower())
    return "_".join(part for part in cleaned.split("_") if part)


def parse_table_titles(workbook) -> dict[int, str]:
    index_sheet_name = None
    for name in workbook.sheetnames:
        if name.upper() == "INDEX":
            index_sheet_name = name
            break
            
    if not index_sheet_name:
        return {}

    titles = {}
    worksheet = workbook[index_sheet_name]
    
    start_row = 3
    for r_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
        if row and len(row) > 0 and row[0] and "Table" in str(row[0]):
            start_row = r_idx + 1
            break
            
    for row in worksheet.iter_rows(min_row=start_row, values_only=True):
        table_cell = row[0] if len(row) > 0 else None
        title = row[1] if len(row) > 1 else None
        if not table_cell or not title:
            continue

        match = re.match(r"Table\s*(\d+)", str(table_cell), re.IGNORECASE)
        if match:
            titles[int(match.group(1))] = str(title).strip()
    return titles


def parse_uploaded_excel(file_bytes: bytes) -> dict:
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    
    tables_sheet_name = None
    for name in workbook.sheetnames:
        if name.upper() == "TABLES":
            tables_sheet_name = name
            break
            
    if not tables_sheet_name:
        raise ValueError("Missing 'Tables' sheet in the workbook.")

    table_titles = parse_table_titles(workbook)
    worksheet = workbook[tables_sheet_name]
    
    rows = list(worksheet.iter_rows(values_only=True))
    parsed_tables = {}
    
    table_starts = []
    for idx, row in enumerate(rows):
        if row and row[0] and str(row[0]).startswith("Table "):
            table_starts.append((row[0], idx))
            
    for i, (table_num_str, start_idx) in enumerate(table_starts):
        match = re.match(r"Table\s*(\d+)", table_num_str, re.IGNORECASE)
        if not match:
            continue
        table_num = int(match.group(1))
        
        title = table_titles.get(table_num, "Unknown")
        if title == "Unknown" and start_idx + 1 < len(rows):
            title = str(rows[start_idx + 1][0] or "Unknown").strip()
            
        header_break_idx = None
        for offset in range(1, 20):
            r_idx = start_idx + offset
            if r_idx >= len(rows):
                break
            if rows[r_idx] and any(str(v).strip().upper() == "TOTAL" for v in rows[r_idx] if v is not None):
                header_break_idx = r_idx
                break
                
        if header_break_idx is None:
            continue
            
        desc_row_idx = header_break_idx - 2
        is_two_row = False
        desc_row = []
        if desc_row_idx > start_idx:
            desc_row = rows[desc_row_idx]
            non_none_desc = [v for v in desc_row[1:] if v is not None and str(v).strip() != ""]
            if len(non_none_desc) > 0:
                is_two_row = True
                
        break_row = rows[header_break_idx]
        
        if is_two_row:
            column_headers = {}
            current_parent = None
            for col_idx in range(1, len(break_row)):
                sub = break_row[col_idx]
                parent = desc_row[col_idx] if col_idx < len(desc_row) else None
                
                if parent is not None and str(parent).strip() != "":
                    current_parent = str(parent).strip()
                    
                if sub is None or str(sub).strip() == "":
                    continue
                    
                sub_str = str(sub).strip()
                if sub_str.upper() == "TOTAL":
                    column_headers[col_idx] = "Total"
                elif sub_str.upper() in ["IPSOS", "KANTAR"]:
                    column_headers[col_idx] = sub_str.upper()
                elif current_parent:
                    column_headers[col_idx] = f"{current_parent}: {sub_str}"
                else:
                    column_headers[col_idx] = sub_str
            data_start_offset = 11
        else:
            column_headers = {}
            for col_idx, val in enumerate(break_row):
                if val not in (None, ""):
                    column_headers[col_idx] = str(val).strip()
            data_start_offset = (header_break_idx - start_idx) + 1
            
        table_data = {}
        row_idx = start_idx + data_start_offset
        
        while row_idx < len(rows):
            row = rows[row_idx]
            row_label = str(row[0]).strip() if row and row[0] is not None else ""
            
            if row_label.startswith("Table ") or row_label == "Sigma":
                break
                
            if row_label in ("", " ") and all(v is None for v in row[1:]):
                row_idx += 1
                continue
                
            row_data = {}
            for col_idx, col_name in column_headers.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    if val not in (None, ""):
                        number = to_number(val)
                        row_data[col_name] = number if number is not None else val
                        
            if row_data:
                table_data[row_label] = row_data
                
            row_idx += 1
            
        if is_two_row:
            for offset in [1, 2]:
                base_row_idx = header_break_idx + offset
                if base_row_idx < len(rows):
                    base_row = rows[base_row_idx]
                    base_label = str(base_row[0]).strip()
                    if "Base" in base_label or "Sample" in base_label:
                        base_data = {}
                        for col_idx, col_name in column_headers.items():
                            if col_idx < len(base_row):
                                val = base_row[col_idx]
                                if val not in (None, ""):
                                    number = to_number(val)
                                    base_data[col_name] = number if number is not None else val
                        if base_data:
                            if "unwtd" in base_label.lower() or "unweighted" in base_label.lower():
                                normalized_label = "Unweighted Sample"
                            elif "wtd" in base_label.lower() or "weighted" in base_label.lower():
                                normalized_label = "Weighted Sample"
                            else:
                                normalized_label = "Unweighted Sample" if "un" in base_label.lower() else "Weighted Sample"
                            table_data[normalized_label] = base_data
                            
        parsed_tables[table_num] = table_data
        
    return {
        str(table_num): {
            "title": table_titles.get(table_num, "Unknown"),
            "data": table_data,
        }
        for table_num, table_data in sorted(parsed_tables.items())
    }



@st.cache_data(show_spinner=False)
def parse_uploaded_excel_cached(file_bytes: bytes) -> dict:
    return parse_uploaded_excel(file_bytes)


def table_label(table_id: str, tables: dict) -> str:
    return str(tables[table_id].get('title', 'Untitled')).strip()


def available_columns(table_info: dict) -> list[str]:
    columns = set()
    for row_values in table_info.get("data", {}).values():
        columns.update(row_values.keys())
    return sorted(column for column in columns if column not in AGENCY_COLUMNS)


def is_response_answer(answer: str) -> bool:
    answer_upper = str(answer).strip().upper()
    normalized_answer = "".join(character for character in answer_upper if character.isalnum() or character.isspace())
    normalized_answer = " ".join(normalized_answer.split())
    is_non_response_row = (
        answer_upper in NON_RESPONSE_ROWS
        or normalized_answer in NON_RESPONSE_ROWS
        or answer_upper.startswith("BASE")
        or normalized_answer.startswith("BASE")
    )
    return not is_non_response_row


def compute_intersection_column(table_info: dict, selected_cols: list[str]) -> str:
    """
    Computes a new intersection column for the given selected columns and
    modifies table_info['data'] in place to include the new column.
    Returns the name of the new column.
    """
    if len(selected_cols) < 2:
        return selected_cols[0] if selected_cols else "Total"
        
    combined_name = " & ".join(selected_cols)
    data = table_info.get("data", {})
    
    # 1. Estimate base sizes for Weighted Sample and Unweighted Sample
    for base_key in ["Weighted Sample", "Unweighted Sample"]:
        if base_key in data:
            base_row = data[base_key]
            total_base = to_number(base_row.get("Total"))
            if total_base is None or total_base <= 0:
                bases = [to_number(base_row.get(col)) for col in selected_cols if col in base_row]
                valid_bases = [b for b in bases if b is not None]
                data[base_key][combined_name] = min(valid_bases) if valid_bases else 0.0
            else:
                bases = [to_number(base_row.get(col)) for col in selected_cols if col in base_row]
                valid_bases = [b for b in bases if b is not None]
                if valid_bases:
                    # Overflow-safe multiplication of ratios instead of high-power exponentiation
                    prod = valid_bases[0]
                    for b in valid_bases[1:]:
                        prod *= (b / total_base)
                    data[base_key][combined_name] = prod
                else:
                    data[base_key][combined_name] = 0.0

    # 2. Compute the cell values for responses
    response_labels = [label for label in data.keys() if is_response_answer(label)]
    
    # Check if the calculated weighted base is 0 (or less than 0.5)
    weighted_base = data.get("Weighted Sample", {}).get(combined_name, 0.0)
    is_empty_base = False
    if weighted_base is not None and weighted_base < 0.5:
        is_empty_base = True
    
    raw_vals = {}
    sum_inputs = {col: 0.0 for col in selected_cols}
    sum_raw = 0.0
    
    for label in response_labels:
        row_vals = data[label]
        total_val = to_number(row_vals.get("Total"))
        
        col_vals = []
        for col in selected_cols:
            val = to_number(row_vals.get(col))
            if val is not None:
                col_vals.append(val)
                sum_inputs[col] += val
                
        if not col_vals:
            continue
            
        if total_val is None or total_val <= 0:
            est = sum(col_vals) / len(col_vals)
        else:
            # Overflow-safe ratio multiplication
            est = col_vals[0]
            for v in col_vals[1:]:
                est *= (v / total_val)
            
        raw_vals[label] = est
        sum_raw += est

    # 3. Normalize percentages
    avg_sum_inputs = sum(sum_inputs.values()) / len(selected_cols) if selected_cols else 100.0
    factor = avg_sum_inputs / sum_raw if sum_raw > 0 else 1.0
    
    for label in response_labels:
        if is_empty_base:
            data[label][combined_name] = 0.0
        elif label in raw_vals:
            data[label][combined_name] = raw_vals[label] * factor
            
    return combined_name


def rows_to_frame(table_info: dict, selected_columns: list[str]) -> pd.DataFrame:
    records = []

    for answer, values in table_info.get("data", {}).items():
        answer_label = str(answer).strip()
        if not is_response_answer(answer_label):
            continue

        answer_label = re.sub(r'\s*\([^)]*specify[^)]*\)', '', answer_label, flags=re.IGNORECASE)
        answer_label = re.sub(r'^\[[^\]]*\]\s*', '', answer_label).strip()

        for top_break in selected_columns:
            if top_break not in values:
                continue

            percentage = to_number(values[top_break])
            if percentage is None:
                continue

            records.append(
                {
                    "Answer": answer_label,
                    "Top Breaks": top_break,
                    "Value": percentage,
                }
            )

    return pd.DataFrame(records)


def top_answers(frame: pd.DataFrame, limit: int) -> pd.DataFrame:
    if frame.empty:
        return frame

    ranking = (
        frame.groupby("Answer", as_index=False)["Value"]
        .max()
        .sort_values("Value", ascending=False)
        .head(limit)
    )
    return frame[frame["Answer"].isin(ranking["Answer"])]


def build_wide_frame(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame

    return (
        frame.pivot_table(index="Answer", columns="Top Breaks", values="Value", aggfunc="first")
        .reset_index()
        .rename_axis(None, axis=1)
    )


def sorted_chart_frame(frame: pd.DataFrame, sort_order: str) -> pd.DataFrame:
    if frame.empty or sort_order == "Original table order":
        return frame

    ascending = sort_order == "Lowest to highest"
    answer_order = (
        frame.groupby("Answer", sort=False)["Value"]
        .max()
        .sort_values(ascending=ascending)
        .index
    )
    return frame.assign(Answer=pd.Categorical(frame["Answer"], categories=answer_order, ordered=True)).sort_values("Answer")


def plot_dashboard_chart(
    chart_type: str,
    chart_frame: pd.DataFrame,
    selected_columns: list[str],
    chart_title: str,
    axis_label: str,
    colors: list[str],
    chart_size: tuple[int, int],
    show_labels: bool,
    label_rotation: int,
    title_position: str,
    legend_position: str,
    show_gridlines: bool,
    pie_label_mode: str,
    round_values: bool = False,
):
    import textwrap
    # Wrap graph title to prevent overflowing chart boundaries
    wrapped_lines = textwrap.wrap(chart_title, width=70)
    wrapped_title = "<br>".join(wrapped_lines)

    width = int(chart_size[0] * 85)
    height = int(chart_size[1] * 85)
    label_format = "%{text:.0f}" if round_values else "%{text:.2f}"

    extra_top_margin = 0
    if len(wrapped_lines) > 1 and not title_position.startswith("Inside"):
        extra_top_margin = (len(wrapped_lines) - 1) * 18
        
    if title_position.startswith("Inside"):
        top_margin = 30
    else:
        top_margin = 55 + extra_top_margin

    if chart_type == "Pie":
        pie_top_break = selected_columns[0]
        pie_frame = chart_frame[chart_frame["Top Breaks"].eq(pie_top_break)]
        if pie_label_mode == "Label + percentage":
            textinfo = "label+percent"
        elif pie_label_mode == "Percentage only":
            textinfo = "percent"
        else:
            textinfo = "label"
            
        fig = px.pie(
            pie_frame,
            names="Answer",
            values="Value",
            color_discrete_sequence=colors,
        )
        fig.update_traces(textinfo=textinfo, textposition="inside" if pie_label_mode != "Label only" else "outside")

    elif chart_type == "Line":
        fig = px.line(
            chart_frame,
            x="Answer",
            y="Value",
            color="Top Breaks",
            color_discrete_sequence=colors,
            markers=True,
            text="Value" if show_labels else None,
        )
        if show_labels:
            fig.update_traces(
                textposition="top right",
                texttemplate=label_format,
            )

    elif chart_type == "Horizontal bar":
        fig = px.bar(
            chart_frame,
            x="Value",
            y="Answer",
            color="Top Breaks",
            orientation="h",
            barmode="group",
            color_discrete_sequence=colors,
            text="Value" if show_labels else None,
        )
        if show_labels:
            fig.update_traces(
                textposition="outside",
                texttemplate=label_format,
            )

    else:  # "Bar"
        fig = px.bar(
            chart_frame,
            x="Answer",
            y="Value",
            color="Top Breaks",
            orientation="v",
            barmode="group",
            color_discrete_sequence=colors,
            text="Value" if show_labels else None,
        )
        if show_labels:
            fig.update_traces(
                textposition="outside",
                texttemplate=label_format,
            )

    # Common layout updates
    fig.update_layout(
        width=width,
        height=height,
        margin=dict(l=50, r=50, t=top_margin, b=50),
        plot_bgcolor="white",
        paper_bgcolor="white",
    )

    if chart_type != "Pie":
        fig.update_xaxes(
            showgrid=show_gridlines,
            gridcolor="#e5e5e5",
            showline=True,
            mirror=True,
            linecolor="#cccccc",
            linewidth=0.5,
        )
        fig.update_yaxes(
            showgrid=show_gridlines,
            gridcolor="#e5e5e5",
            showline=True,
            mirror=True,
            linecolor="#cccccc",
            linewidth=0.5,
        )
        if chart_type == "Horizontal bar":
            fig.update_xaxes(title_text=axis_label)
            fig.update_yaxes(title_text="")
        else:
            fig.update_xaxes(title_text="", tickangle=label_rotation)
            fig.update_yaxes(title_text=axis_label)

    # Position Title
    title_positions = {
        "Above chart": dict(xref="paper", yref="container", x=0.5, y=0.98, xanchor="center", yanchor="top"),
        "Inside top left": dict(xref="paper", yref="paper", x=0.02, y=0.95, xanchor="left", yanchor="top"),
        "Inside top center": dict(xref="paper", yref="paper", x=0.5, y=0.95, xanchor="center", yanchor="top"),
        "Inside top right": dict(xref="paper", yref="paper", x=0.98, y=0.95, xanchor="right", yanchor="top"),
        "Inside bottom left": dict(xref="paper", yref="paper", x=0.02, y=0.05, xanchor="left", yanchor="bottom"),
        "Inside bottom right": dict(xref="paper", yref="paper", x=0.98, y=0.05, xanchor="right", yanchor="bottom"),
    }
    t_pos = title_positions.get(title_position, title_positions["Above chart"])
    
    if title_position.startswith("Inside"):
        title_text = f'<b><span style="background-color: rgba(255,255,255,0.75); padding: 3px;">{wrapped_title}</span></b>'
    else:
        title_text = f'<b>{wrapped_title}</b>'

    fig.update_layout(
        title=dict(
            text=title_text,
            font=dict(size=16, color="#333333"),
            **t_pos
        )
    )

    # Position Legend
    if legend_position == "Hidden":
        fig.update_layout(showlegend=False)
    else:
        legend_configs = {
            "Outside right": dict(x=1.02, y=0.5, xanchor="left", yanchor="middle", orientation="v"),
            "Outside top": dict(x=0.5, y=1.15, xanchor="center", yanchor="bottom", orientation="h"),
            "Outside bottom": dict(x=0.5, y=-0.25, xanchor="center", yanchor="top", orientation="h"),
            "Inside top left": dict(x=0.02, y=0.98, xanchor="left", yanchor="top", orientation="v"),
            "Inside top right": dict(x=0.98, y=0.98, xanchor="right", yanchor="top", orientation="v"),
            "Inside bottom left": dict(x=0.02, y=0.02, xanchor="left", yanchor="bottom", orientation="v"),
            "Inside bottom right": dict(x=0.98, y=0.02, xanchor="right", yanchor="bottom", orientation="v"),
        }
        l_config = legend_configs.get(legend_position, legend_configs["Outside right"])
        fig.update_layout(
            showlegend=True,
            legend=dict(
                title=dict(text="Legend"),
                bgcolor="rgba(255,255,255,0.8)",
                bordercolor="#e5e5e5",
                borderwidth=0.5,
                **l_config
            )
        )

    # Wrap legend text for long top break names
    for trace in fig.data:
        if trace.name:
            trace.name = "<br>".join(textwrap.wrap(trace.name, width=30))

    return fig

st.markdown(
    """
    <style>
    .block-container { padding-top: 1.5rem; }
    div[data-testid="stMetric"] {
        border: 1px solid #e6e8eb;
        border-radius: 8px;
        padding: 12px 14px;
        background: #ffffff;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("Maruti Survey Dashboard")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xlsm"])
if uploaded_file is None:
    st.session_state["last_file_bytes"] = None
    st.info("Upload the summary Excel file to begin.")
    st.stop()

if "last_file_bytes" not in st.session_state:
    st.session_state["last_file_bytes"] = None

file_bytes = uploaded_file.getvalue()
if st.session_state["last_file_bytes"] != file_bytes:
    st.cache_data.clear()
    st.session_state["last_file_bytes"] = file_bytes

try:
    tables = parse_uploaded_excel_cached(file_bytes)
except Exception as exc:
    st.error(f"Could not parse this workbook: {exc}")
    st.stop()

if not tables:
    st.error("No tables were found in the uploaded workbook.")
    st.stop()


table_ids = sorted(tables.keys(), key=lambda table_id: int(table_id))

with st.sidebar:
    st.header("Dashboard Controls")
    st.success(f"Parsed {len(table_ids)} tables")

    filtered_ids = [
        table_id
        for table_id in table_ids
    ]

    if not filtered_ids:
        st.warning("No questions match your search.")
        st.stop()

    selected_table_id = st.selectbox(
        "Question",
        options=filtered_ids,
        format_func=lambda table_id: table_label(table_id, tables),
    )

    selected_table = tables[selected_table_id]
    columns_in_table = available_columns(selected_table)
    default_columns = [column for column in ["Total"] if column in columns_in_table]
    
    # State initialization for dynamic groups
    if "group_keys" not in st.session_state:
        st.session_state["group_keys"] = [0]
    if "next_group_key" not in st.session_state:
        st.session_state["next_group_key"] = 1

    active_columns = []
    keys_to_remove = []
    
    st.markdown("### Top Breaks Comparison")
    for idx, key in enumerate(st.session_state["group_keys"]):
        st.markdown(f"**Group {idx + 1}**")
        
        # Determine defaults for group
        if idx == 0:
            group_default = [col for col in ["Total"] if col in columns_in_table]
            if not group_default and columns_in_table:
                group_default = [columns_in_table[0]]
        else:
            group_default = [columns_in_table[0]] if columns_in_table else []
            
        group_cols = st.multiselect(
            f"Select column(s) for Group {idx + 1}",
            options=columns_in_table,
            default=group_default,
            key=f"group_select_{key}"
        )
        
        # Only show a delete button if there's more than one group
        if len(st.session_state["group_keys"]) > 1:
            if st.button("🗑️ Remove Group", key=f"remove_btn_{key}"):
                keys_to_remove.append(key)
                
        if group_cols:
            if len(group_cols) >= 2:
                combined_col = compute_intersection_column(selected_table, group_cols)
                active_columns.append(combined_col)
            else:
                active_columns.append(group_cols[0])
                
        st.markdown("---")
        
    # Handle group deletions
    if keys_to_remove:
        for k in keys_to_remove:
            st.session_state["group_keys"].remove(k)
        st.rerun()
        
    # Add new group button
    if st.button("+ Add Top Break Group"):
        st.session_state["group_keys"].append(st.session_state["next_group_key"])
        st.session_state["next_group_key"] += 1
        st.rerun()

    selected_columns = active_columns
    selected_group = ", ".join(active_columns) if active_columns else "None"

    chart_type = st.selectbox("Chart", ["Bar", "Horizontal bar", "Line", "Pie"])
    show_all = st.checkbox("Show all side breaks", value=False)
    if not show_all:
        top_n = st.slider("Show top side breaks", min_value=5, max_value=50, value=20, step=5)
    else:
        top_n = None
    sort_order = st.selectbox("Sort order", ["Highest to lowest", "Lowest to highest", "Original table order"])

    st.subheader("Chart Customization")
    default_title = f"{clean_title(selected_table.get('title', 'Untitled'))} by {selected_group}"
    chart_title = st.text_input("Chart title", value=default_title)
    axis_label = st.text_input("Axis label", value="Percentage")
    palette_name = st.selectbox("Color palette", list(COLOR_PALETTES.keys()))
    chart_size_name = st.selectbox("Chart size", list(CHART_SIZES.keys()), index=1)
    show_labels = st.checkbox("Show data labels", value=False)
    round_values = st.checkbox("Round off percentages", value=False)
    show_gridlines = st.checkbox("Show gridlines", value=True)
    label_rotation = st.slider("X label rotation", min_value=0, max_value=90, value=35, step=5)
    title_position = st.selectbox(
        "Title position",
        ["Above chart", "Inside top left", "Inside top center", "Inside top right", "Inside bottom left", "Inside bottom right"],
    )
    legend_position = st.selectbox(
        "Legend position",
        [
            "Outside right",
            "Outside top",
            "Outside bottom",
            "Inside top left",
            "Inside top right",
            "Inside bottom left",
            "Inside bottom right",
            "Hidden",
        ],
    )
    pie_label_mode = st.selectbox("Pie labels", ["Label + percentage", "Percentage only", "Label only"])

    st.markdown("---")

if not selected_columns:
    st.error("Choose at least one top break for this question.")
    st.stop()

raw_frame = rows_to_frame(selected_table, active_columns)
if raw_frame.empty:
    st.warning("This question has no numeric data for the selected top breaks.")
    st.stop()

if round_values:
    raw_frame["Value"] = raw_frame["Value"].round(0).astype(int)

chart_frame_data = raw_frame[
    (raw_frame["Answer"].str.lower().str.strip() != "unspecified") &
    (~raw_frame["Answer"].str.contains(r'(top\s*2\s*box|bottom\s*2\s*box|top\s*two\s*box|bottom\s*two\s*box|t2b|b2b)', case=False, regex=True))
]
if show_all:
    chart_frame = sorted_chart_frame(chart_frame_data, sort_order)
else:
    chart_frame = sorted_chart_frame(top_answers(chart_frame_data, top_n), sort_order)
wide_frame = build_wide_frame(raw_frame)

st.subheader(clean_title(table_label(selected_table_id, tables)))
metric_1, metric_2 = st.columns(2)

unique_answers = raw_frame["Answer"].unique()
has_unspecified = any(str(ans).lower().strip() == "unspecified" for ans in unique_answers)
side_breaks_count = len(unique_answers) - 1 if has_unspecified else len(unique_answers)
metric_1.metric("Side Breaks", side_breaks_count)
metric_2.metric("Groups", len(selected_columns))

# Extract correct sample base from weighted sample row based on selected columns/intersection
unique_active_cols = []
for col in active_columns:
    if col not in unique_active_cols:
        unique_active_cols.append(col)

base_records = []
for i, col in enumerate(unique_active_cols):
    w_base = selected_table.get("data", {}).get("Weighted Sample", {}).get(col, "NA")
    u_base = selected_table.get("data", {}).get("Unweighted Sample", {}).get(col, "NA")
    
    w_str = f"{w_base:,.0f}" if isinstance(w_base, (int, float)) else str(w_base)
    u_str = f"{u_base:,.0f}" if isinstance(u_base, (int, float)) else str(u_base)
    
    base_records.append({
        "Group": f"Group {i+1}",
        "Top Break Selection": col,
        "Weighted Sample Base": w_str,
        "Unweighted Sample Base": u_str
    })
base_df = pd.DataFrame(base_records)

st.markdown("##### Sample Bases")
st.dataframe(base_df, hide_index=True, use_container_width=True)

if chart_type == "Pie" and len(active_columns) > 1:
    st.info(f"Pie chart is showing {active_columns[0]}. Choose only one top break for a different pie.")

fig = plot_dashboard_chart(
    chart_type=chart_type,
    chart_frame=chart_frame,
    selected_columns=active_columns,
    chart_title=chart_title,
    axis_label=axis_label,
    colors=COLOR_PALETTES[palette_name],
    chart_size=CHART_SIZES[chart_size_name],
    show_labels=show_labels,
    label_rotation=label_rotation,
    title_position=title_position,
    legend_position=legend_position,
    show_gridlines=show_gridlines,
    pie_label_mode=pie_label_mode,
    round_values=round_values,
)
download_filename = f"table_{selected_table_id}_{safe_filename(chart_type)}_{safe_filename(selected_group)}"
st.plotly_chart(
    fig,
    use_container_width=True,
    config={"toImageButtonOptions": {"filename": download_filename}}
)

st.divider()
st.subheader("Data")
sort_column = "Total" if "Total" in wide_frame.columns else active_columns[0]
if sort_column in wide_frame.columns:
    wide_frame = wide_frame.sort_values(sort_column, ascending=False)
st.dataframe(wide_frame.rename(columns={"Answer": "Side Breaks"}), hide_index=True, use_container_width=True)
